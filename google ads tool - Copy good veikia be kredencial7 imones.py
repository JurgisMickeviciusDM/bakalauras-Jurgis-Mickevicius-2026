# BatchService šablonas - originalus šablonas su žymėjimais pakeitimui#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import time
import json
import sys
import os
import random
import base64
import requests
import urllib.parse
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment

# —————————————————————————————————————————————————————————————————————————
#  KONFIGŪRACIJOS PARAMETRAI
chrome_profile_path = r"C:\selenium_uc_profile"  # Chrome profilio aplanko kelias
KP_HOME = (
    "https://ads.google.com/aw/keywordplanner/home"
    "*****************************"
    "****************************"
)
# laikinas raktažodis pirminei inicializacijai
INITIAL_KEYWORD = "SEO"
DEBUG = False  # True — detalesnių žurnalų gavimui
JSON_OUTPUT = True  # True — JSON formatu, False — įprasta forma

# Konfigūracijos kintamieji TEMPLATE_DATA šablonui
ACCOUNT_ID = "************"
USER_ID = "**************"
CLIENT_ID = "*************"
SID = "-*******************"
CURRENCY_CODE = "EUR"
LANGUAGE_CODE = "1029"
LOCATION_CODE = "2440"
NETWORK_CODE = "1"
DATE_START_YEAR = 2022
DATE_START_MONTH = 1
DATE_START_DAY = 1
DATE_END_YEAR = 2025
DATE_END_MONTH = 11
DATE_END_DAY = 30
RESULTS_LIMIT_START = 0
RESULTS_LIMIT_COUNT = 500
TEMPLATE_DATA = r"""hl=en_US&__lu=260540043&__u=USER_ID_PLACEHOLDER&__c=CLIENT_ID_PLACEHOLDER&f.sid=SID_PLACEHOLDER&ps=aw&__ar={
"2":["{\"1\":{\"3\":{\"1\":\"ACCOUNT_ID_PLACEHOLDER\"},\"5\":\"TABLE\"},\"2\":{\"1\":[\"text\",\"search_volume\",\"search_volume_trends\",\"recent_search_trend_change\",\"competition\",\"ad_impression_share\",\"recent_yoy_search_trend_change\",\"account_status\",\"is_in_plan\",\"is_in_account\",\"is_negative\",\"bid_min\",\"bid_max\",\"competition_index\",\"organic_impression_share\",\"organic_average_position\",\"keyword_variants\"],\"2\":[{\"1\":\"is_adult_idea\",\"2\":1,\"4\":[{\"1\":true}]},{\"1\":\"keyword_seed\",\"2\":1,\"4\":[{\"6\":\"KEYWORD_PLACEHOLDER\"}]},{\"1\":\"currency_code\",\"2\":1,\"4\":[{\"6\":\"CURRENCY_CODE_PLACEHOLDER\"}]},{\"1\":\"language\",\"2\":1,\"4\":[{\"3\":\"LANGUAGE_CODE_PLACEHOLDER\"}]},{\"1\":\"locations\",\"4\":[{\"3\":\"LOCATION_CODE_PLACEHOLDER\"}]},{\"1\":\"network\",\"2\":1,\"4\":[{\"3\":\"NETWORK_CODE_PLACEHOLDER\"}]},{\"1\":\"search_volume_types\",\"2\":1,\"4\":[{\"3\":\"4\"},{\"3\":\"1\"},{\"3\":\"3\"},{\"3\":\"5\"}]},{\"1\":\"location_segmentation\",\"2\":1,\"4\":[{\"3\":\"2\"}]},{\"1\":\"skip_location_chart\",\"2\":1,\"4\":[{\"1\":true}]}],\"3\":[{\"1\":\"text\",\"2\":2}],\"4\":{\"1\":{\"1\":DATE_START_YEAR_PLACEHOLDER,\"2\":DATE_START_MONTH_PLACEHOLDER,\"3\":DATE_START_DAY_PLACEHOLDER},\"2\":{\"1\":DATE_END_YEAR_PLACEHOLDER,\"2\":DATE_END_MONTH_PLACEHOLDER,\"3\":DATE_END_DAY_PLACEHOLDER}},\"7\":{\"1\":RESULTS_LIMIT_START_PLACEHOLDER,\"2\":RESULTS_LIMIT_COUNT_PLACEHOLDER},\"14\":true},\"3\":[{\"1\":\"AWN_KP_USE_CONTRA_LLM\",\"2\":\"TRUE\"},{\"1\":\"AWN_KP_USE_CONTRA_REPLACEMENT_SYNONYMS\",\"2\":\"TRUE\"},{\"1\":\"ADD_CHARTS_REQUEST\",\"2\":\"TRUE\"},{\"1\":\"INVERT_GROUPING_ENTITIES\",\"2\":\"TRUE\"},{\"1\":\"PUSH_ID_FILTERS_INTO_LABEL_SUBVIEW\",\"2\":\"TRUE\"}]}","{\"1\":{\"3\":{\"1\":\"ACCOUNT_ID_PLACEHOLDER\"},\"5\":\"TABLE_CHART\"},\"2\":{\"2\":[{\"1\":\"is_adult_idea\",\"2\":1,\"4\":[{\"1\":true}]},{\"1\":\"keyword_seed\",\"2\":1,\"4\":[{\"6\":\"KEYWORD_PLACEHOLDER\"}]},{\"1\":\"currency_code\",\"2\":1,\"4\":[{\"6\":\"CURRENCY_CODE_PLACEHOLDER\"}]},{\"1\":\"language\",\"2\":1,\"4\":[{\"3\":\"LANGUAGE_CODE_PLACEHOLDER\"}]},{\"1\":\"locations\",\"4\":[{\"3\":\"LOCATION_CODE_PLACEHOLDER\"}]},{\"1\":\"network\",\"2\":1,\"4\":[{\"3\":\"NETWORK_CODE_PLACEHOLDER\"}]},{\"1\":\"search_volume_types\",\"2\":1,\"4\":[{\"3\":\"4\"},{\"3\":\"1\"},{\"3\":\"3\"},{\"3\":\"5\"}]},{\"1\":\"location_segmentation\",\"2\":1,\"4\":[{\"3\":\"2\"}]},{\"1\":\"skip_location_chart\",\"2\":1,\"4\":[{\"1\":true}]}],\"4\":{\"1\":{\"1\":DATE_START_YEAR_PLACEHOLDER,\"2\":DATE_START_MONTH_PLACEHOLDER,\"3\":DATE_START_DAY_PLACEHOLDER},\"2\":{\"1\":DATE_END_YEAR_PLACEHOLDER,\"2\":DATE_END_MONTH_PLACEHOLDER,\"3\":DATE_END_DAY_PLACEHOLDER}},\"14\":true},\"3\":[{\"1\":\"AWN_KP_USE_CONTRA_LLM\",\"2\":\"TRUE\"},{\"1\":\"AWN_KP_USE_CONTRA_REPLACEMENT_SYNONYMS\",\"2\":\"TRUE\"},{\"1\":\"ADD_CHARTS_REQUEST\",\"2\":\"TRUE\"},{\"1\":\"INVERT_GROUPING_ENTITIES\",\"2\":\"TRUE\"},{\"1\":\"PUSH_ID_FILTERS_INTO_LABEL_SUBVIEW\",\"2\":\"TRUE\"}]}"],"3":[{"1":"ads.awapps.anji.proto.kp.IdeasService","2":"List"},{"1":"ads.awapps.anji.proto.kp.IdeasService","2":"Charts"}]}&activityContext=IdeasHeader.SearchBar.GetSearchResults&requestPriority=HIGH_LATENCY_SENSITIVE&activityType=INTERACTIVE&activityId=2095505314844441&uniqueFingerprint=SID_PLACEHOLDER_2095505314844441_1&previousPlace=/aw/keywordplanner/ideas/new&activityName=IdeasHeader.SearchBar.GetSearchResults&destinationPlace=/aw/keywordplanner/ideas/new"""



def human_type(element, text, min_delay=0.1, max_delay=0.3):
    """Emuliuoja teksto įvedimą simbolis po simbolio"""
    for ch in text:
        element.send_keys(ch)
        time.sleep(random.uniform(min_delay, max_delay))


def extract_params(initial_keyword: str):
    """
    Naudojant Selenium:
      1) pereina į Keyword Planner,
      2) įveda initial_keyword,
      3) paspaudžia Enter ir spragsti Get results,
      4) perima POST į BatchService/Batch,
    ir grąžina (batch_url, originalūs antraštės, cookie_header).
    """
    options = uc.ChromeOptions()
    options.add_argument(f"--user-data-dir={chrome_profile_path}")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.set_capability("goog:loggingPrefs", {"performance": "ALL"})
    options.add_experimental_option("perfLoggingPrefs", {"enableNetwork": True})
    driver = uc.Chrome(options=options, version_main=144)
    wait = WebDriverWait(driver, 90)
    try:
        driver.get(KP_HOME)
        time.sleep(5)
        # 1. Discover new keywords
        btn = wait.until(EC.element_to_be_clickable(
            (By.XPATH, "//span[normalize-space(text())='Discover new keywords']")))
        time.sleep(1);
        driver.execute_script("arguments[0].click();", btn)
        # 2. Įvesties laukas
        inp = wait.until(EC.visibility_of_element_located(
            (By.CSS_SELECTOR, "multi-suggest-input input.search-input")))
        inp.clear()
        human_type(inp, initial_keyword, 0.15, 0.35)
        time.sleep(0.5)
        # 3. Siųsti Enter
        driver.get_log("performance")  # išvalyti senus žurnalus
        inp.send_keys(Keys.ENTER)
        time.sleep(1)
        # 4. Spragtelėti Get results
        get_btn = wait.until(EC.presence_of_element_located(
            (By.XPATH,
             "//material-button[contains(@class,'submit-button') and .//div[normalize-space(text())='Get results']]")))
        wait.until(lambda d: get_btn.get_attribute("aria-disabled") == "false")
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", get_btn)
        time.sleep(0.3)
        driver.execute_script("arguments[0].click();", get_btn)
        # 5. POST perimimas
        batch_url = None
        orig_headers = {}
        for _ in range(30):
            time.sleep(1)
            for e in driver.get_log("performance"):
                m = json.loads(e["message"])["message"]
                if m.get("method") != "Network.requestWillBeSent":
                    continue
                req = m["params"]["request"]
                url = req.get("url", "")
                if req.get("method") == "POST" and "BatchService/Batch" in url:
                    batch_url = url
                    orig_headers = req.get("headers", {})
                    break
            if batch_url:
                break
        if not batch_url:
            print("Nepavyko perimti POST į BatchService/Batch")
            sys.exit(1)
        # Renkame slapuką
        cookie_header = "; ".join(f"{c['name']}={c['value']}" for c in driver.get_cookies())
        parsed = urllib.parse.urlparse(batch_url)
        qs = urllib.parse.parse_qs(parsed.query)
        
        # Ištraukiame parametrus (jei nėra - grąžins None arba tuščią)
        # __u -> USER_ID
        # __c -> CLIENT_ID
        # f.sid -> SID
        
        got_uid = qs.get("__u", [""])[0]
        got_cid = qs.get("__c", [""])[0]
        got_sid = qs.get("f.sid", [""])[0]
        
        print(f"🔍 Aptikti parametrai (Batch): UserID={got_uid}, ClientID={got_cid}, SID={got_sid}")

        # Papildomas patikrinimas iš URL (nes Batch dažnai neturi __u / __c, ir beveik niekada OCID)
        current_url = driver.current_url
        print(f"Dabartinis URL: {current_url}")
        parsed_current = urllib.parse.urlparse(current_url)
        qs_current = urllib.parse.parse_qs(parsed_current.query)

        url_uid = qs_current.get("__u", [""])[0]
        url_cid = qs_current.get("__c", [""])[0]
        url_ocid = qs_current.get("ocid", [""])[0]
        
        print(f"Aptikti parametrai (URL): UserID={url_uid}, ClientID={url_cid}, OCID={url_ocid}")

        # Prioritetas: URL > Batch (nes URL tikslesnis dabartiniam vaizdui)
        final_uid = url_uid if url_uid else got_uid
        final_cid = url_cid if url_cid else got_cid
        final_sid = got_sid # SID būna tik Batch užklausoje
        final_acc = url_ocid # Account ID būna URL (ocid)

        return batch_url, orig_headers, cookie_header, final_uid, final_cid, final_sid, final_acc
    except Exception as e:
        print(f"Klaida vykdant Selenium veiksmus: {e}")
        driver.save_screenshot("debug_error.png")
        print("Įvyko klaida! Išsaugotas ekranvaizdis: debug_error.png. Patikrinkite šį failą, kad pamatytumėte, kas matoma naršyklėje.")
        raise e
    finally:
        driver.quit()


def make_request(keyword: str, batch_url: str, orig_headers: dict, cookie_header: str):
    # Tiesiog VISADA naudojame JSON - ignoruojame visus kintamuosius
    data = TEMPLATE_DATA.replace("USER_ID_PLACEHOLDER", str(USER_ID)) \
        .replace("CLIENT_ID_PLACEHOLDER", str(CLIENT_ID)) \
        .replace("SID_PLACEHOLDER", str(SID)) \
        .replace("ACCOUNT_ID_PLACEHOLDER", str(ACCOUNT_ID)) \
        .replace("CURRENCY_CODE_PLACEHOLDER", str(CURRENCY_CODE)) \
        .replace("LANGUAGE_CODE_PLACEHOLDER", str(LANGUAGE_CODE)) \
        .replace("LOCATION_CODE_PLACEHOLDER", str(LOCATION_CODE)) \
        .replace("NETWORK_CODE_PLACEHOLDER", str(NETWORK_CODE)) \
        .replace("DATE_START_YEAR_PLACEHOLDER", str(DATE_START_YEAR)) \
        .replace("DATE_START_MONTH_PLACEHOLDER", str(DATE_START_MONTH)) \
        .replace("DATE_START_DAY_PLACEHOLDER", str(DATE_START_DAY)) \
        .replace("DATE_END_YEAR_PLACEHOLDER", str(DATE_END_YEAR)) \
        .replace("DATE_END_MONTH_PLACEHOLDER", str(DATE_END_MONTH)) \
        .replace("DATE_END_DAY_PLACEHOLDER", str(DATE_END_DAY)) \
        .replace("RESULTS_LIMIT_START_PLACEHOLDER", str(RESULTS_LIMIT_START)) \
        .replace("RESULTS_LIMIT_COUNT_PLACEHOLDER", str(RESULTS_LIMIT_COUNT)) \
        .replace("KEYWORD_PLACEHOLDER", keyword)

    quoted = urllib.parse.quote(data, safe=":/?&=")
    encoded = base64.b64encode(quoted.encode())
    payload = base64.b64decode(encoded)
    content_length = str(len(encoded))

    headers = orig_headers.copy()
    headers.update({
        "content-length": content_length,
        "content-type": "application/x-www-form-urlencoded",
        "cookie": cookie_header,
        "origin": "https://ads.google.com",
        "referer": KP_HOME,
    })

    resp = requests.post(batch_url, headers=headers, data=payload, timeout=30)
    resp.raise_for_status()
    body = resp.json()

    # Funkcija datoms generuoti pagal konfigūracijos kintamuosius
    def generate_dates():
        dates = []
        year = DATE_START_YEAR
        month = DATE_START_MONTH

        end_year = DATE_END_YEAR
        end_month = DATE_END_MONTH

        while year < end_year or (year == end_year and month <= end_month):
            dates.append(f"{year}-{month:02d}")
            month += 1
            if month > 12:
                month = 1
                year += 1

        return dates

    # Funkcija mikroeurų konvertavimui į eurus
    def micros_to_euros(micros_str):
        if not micros_str or micros_str == "":
            return "0.00"
        try:
            micros = int(micros_str)
            euros = micros / 1000000  # Google Ads naudoja mikroeurus
            return f"{euros:.2f}"
        except:
            return "0.00"

    # TRUMPAS JSON formatas
    if "2" in body and body["2"]:
        kw_data = json.loads(body["2"][0])
        suggestions = kw_data.get("2", {}).get("1", [])

        date_labels = generate_dates()

        result = {
            "keyword": keyword,
            "total": len(suggestions),
            "date_range": f"{DATE_START_YEAR}-{DATE_START_MONTH:02d} iki {DATE_END_YEAR}-{DATE_END_MONTH:02d}",
            "data": []
        }

        for kw in suggestions:
            monthly_searches = kw.get("1021", [])

            monthly_data = {}
            for i, volume in enumerate(monthly_searches):
                if i < len(date_labels):
                    monthly_data[date_labels[i]] = volume

            suggestion = {
                "text": kw.get("1", ""),
                "volume": kw.get("1000", ""),
                "competition": kw.get("1001", ""),
                "min_bid_euros": micros_to_euros(kw.get("1011", "")),
                "max_bid_euros": micros_to_euros(kw.get("1012", "")),
                "monthly_data": monthly_data,
                "trends": {
                    "yoy_change": kw.get("1035", ""),
                    "recent_change": kw.get("1037", "")
                }
            }
            result["data"].append(suggestion)

        print(json.dumps(result, indent=2, ensure_ascii=False))
        save_results_to_file(result)
        save_results_to_excel(result)


def save_results_to_excel(new_data, filename="Keyword_Stats.xlsx"):
    try:
        # Prepare data for DataFrame
        rows = []
        keyword = new_data.get("keyword")
        
        # Get date range string from config or data
        # Note: The user want "December 1, 2024 - November 30, 2025" format
        # We constructed date_range string in make_request but it was YYYY-MM formatted.
        # Let's reconstruct consistent with the user example if possible, or use the one we have.
        # User example: "December 1, 2024 - November 30, 2025"
        start_date = datetime(DATE_START_YEAR, DATE_START_MONTH, DATE_START_DAY)
        end_date = datetime(DATE_END_YEAR, DATE_END_MONTH, DATE_END_DAY)
        date_range_str = f"{start_date.strftime('%B %d, %Y')} - {end_date.strftime('%B %d, %Y')}"

        # Generate dynamic headers for searches
        # We need to ensure we have the same months as in the data.
        # The data['monthly_data'] keys are "YYYY-MM".
        # We want "Searches: Mon YYYY"
        
        for item in new_data.get("data", []):
            row = {
                "Keyword": item.get("text"),
                "Currency": CURRENCY_CODE, # Global var
                "Avg. monthly searches": item.get("volume"),
                "Three month change": item["trends"].get("recent_change"),
                "YoY change": item["trends"].get("yoy_change"),
                "Competition": item.get("competition"),
                "Competition (indexed value)": "", # JSON '1002' not mapped in current script, leaving empty or need to map if available
                "Top of page bid (low range)": item.get("min_bid_euros"),
                "Top of page bid (high range)": item.get("max_bid_euros"),
                "Ad impression share": "", # Not extracted
                "Organic impression share": "", # Not extracted
                "Organic average position": "", # Not extracted
                "In account?": "", # Not extracted
                "In plan?": "", # Not extracted
            }
            
            # Add monthly search data
            monthly_data = item.get("monthly_data", {})
            for date_key, volume in monthly_data.items():
                # date_key is YYYY-MM
                try:
                    dt = datetime.strptime(date_key, "%Y-%m")
                    header_name = f"Searches: {dt.strftime('%b %Y')}"
                    row[header_name] = volume
                except:
                    pass
            
            rows.append(row)

        df = pd.DataFrame(rows)
        
        # Define column order (base columns + dynamic monthly columns)
        base_columns = [
            "Keyword", "Currency", "Avg. monthly searches", "Three month change", "YoY change",
            "Competition", "Competition (indexed value)", "Top of page bid (low range)", 
            "Top of page bid (high range)", "Ad impression share", "Organic impression share", 
            "Organic average position", "In account?", "In plan?"
        ]
        
        # Identify monthly columns present in this dataframe
        monthly_cols = [c for c in df.columns if c.startswith("Searches: ")]
        # Sort monthly columns chronologically? They usually come in order from the API/dict, but good to ensure.
        # Simple string sort might fail for "Jan 2025" vs "Dec 2024".
        # Let's rely on the order they were inserted which depends on generate_dates() order.
        
        final_columns = base_columns + monthly_cols
        
        # Reorder df, adding missing columns as empty
        for col in base_columns:
            if col not in df.columns:
                df[col] = ""
                
        df = df[final_columns]

        # Check if file exists to determine if we write header
        file_exists = os.path.exists(filename)
        
        if not file_exists:
            # Create new workbook with custom header rows
            wb = openpyxl.Workbook()
            ws = wb.active
            
            # Row 1: Title
            timestamp = datetime.now().strftime("%Y-%m-%d at %H_%M_%S")
            title_cell = ws.cell(row=1, column=1, value=f"Keyword Stats {timestamp}")
            
            # Row 2: Date Range
            ws.cell(row=2, column=1, value=date_range_str)
            
            # Row 3: Headers
            for col_idx, column_name in enumerate(final_columns, 1):
                cell = ws.cell(row=3, column=col_idx, value=column_name)
                cell.font = Font(bold=True)
            
            # Append data starting from row 4
            for r in dataframe_to_rows(df, index=False, header=False):
                ws.append(r)
            
            wb.save(filename)
            print(f"Sukurtas naujas Excel failas: {filename}")
            
        else:
            # Append to existing file
            # Load workbook
            wb = openpyxl.load_workbook(filename)
            ws = wb.active
            
            # Find the last row
            # We assume the structure is fixed.
            # We need to match columns of the new data with the existing file.
            # Read existing headers from row 3
            existing_headers = [cell.value for cell in ws[3]]
            
            # Align new df to existing headers
            # If new columns appeared (e.g. date range changed), this might be tricky. 
            # Assuming consistent date range for the session.
            
            aligned_rows = []
            for _, row in df.iterrows():
                new_row = []
                for header in existing_headers:
                    if header in row:
                        new_row.append(row[header])
                    else:
                        new_row.append("")
                aligned_rows.append(new_row)
                
            for r in aligned_rows:
                ws.append(r)
                
            wb.save(filename)
            print(f"Duomenys papildyti faile: {filename}")

    except Exception as e:
        print(f"Klaida saugant Excel: {e}")

def save_results_to_file(data, filename="google_ads_data.json"):
    try:
        if os.path.exists(filename):
            with open(filename, 'r', encoding='utf-8') as f:
                try:
                    current_data = json.load(f)
                    if not isinstance(current_data, list):
                        current_data = [] 
                except json.JSONDecodeError:
                    current_data = [] # Handle empty or corrupt file
        else:
            current_data = []
        
        current_data.append(data)
        
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(current_data, f, indent=2, ensure_ascii=False)
        print(f"Duomenys išsaugoti į {filename}")
    except Exception as e:
        print(f"Klaida saugant duomenis: {e}")
def main():
    # PRIVERSTINAI NUSTATOME JSON FORMATĄ
    global JSON_OUTPUT
    JSON_OUTPUT = True
    print(f"JSON_OUTPUT nustatytas į: {JSON_OUTPUT}")

    # 1) Pereimame URL ir originaliąsias antraštes vieną kartą
    print("Inicializacija... ištraukiame užklausos parametrus per Selenium.")
    # Atnaujiname globals su ištraukta info
    batch_url, orig_headers, cookie_header, new_uid, new_cid, new_sid, new_acc = extract_params(INITIAL_KEYWORD)
    
    global USER_ID, CLIENT_ID, SID, ACCOUNT_ID
    if new_uid: USER_ID = new_uid
    if new_cid: CLIENT_ID = new_cid
    if new_sid: SID = new_sid
    if new_acc: ACCOUNT_ID = new_acc
    
    print(f"Nustatyta sesija: UID={USER_ID}, CID={CLIENT_ID}, SID={SID}, ACC={ACCOUNT_ID}")
    print("Baigta! Dabar įveskite raktažodžius (exit išeiti).")

    # 2) Interaktyvus ciklas
    while True:
        kw = input("\nĮveskite raktažodį (arba 'exit'): ").strip()
        if not kw or kw.lower() in ("exit", "quit"):
            print("Baigiamas darbas.")
            break
        try:
            make_request(kw, batch_url, orig_headers, cookie_header)
        except Exception as e:
            print("Klaida užklausos metu:", e)


if __name__ == "__main__":
    main()